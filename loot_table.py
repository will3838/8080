from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List

from openpyxl import load_workbook


@dataclass(frozen=True)
class Item:
    item_id: str
    name: str
    weight: float
    icon_path: Path


@dataclass
class LootTable:
    items: List[Item]
    prefix_sums: List[float]
    sum_weights: float

    def choose_by_r(self, r: float) -> Item:
        """Pick item by random value r in [0, 1)."""
        if not 0 <= r < 1:
            raise ValueError("r must be in [0, 1)")

        target = r * self.sum_weights
        lo, hi = 0, len(self.prefix_sums) - 1
        while lo < hi:
            mid = (lo + hi) // 2
            if target < self.prefix_sums[mid]:
                hi = mid
            else:
                lo = mid + 1
        return self.items[lo]


def _normalize_id(raw_id: object) -> str:
    if raw_id is None:
        return ""
    if isinstance(raw_id, float) and raw_id.is_integer():
        return str(int(raw_id))
    return str(raw_id).strip()


def load_loot_table(excel_path: Path, icons_dir: Path) -> LootTable:
    if not excel_path.exists():
        raise ValueError(f"Excel file is missing: {excel_path}")

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    items: List[Item] = []
    seen_ids: set[str] = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        raw_id = row[0] if len(row) > 0 else None
        raw_name = row[1] if len(row) > 1 else None
        raw_chance = row[2] if len(row) > 2 else None

        if raw_id is None and raw_name is None and raw_chance is None:
            continue

        item_id = _normalize_id(raw_id)
        name = str(raw_name).strip() if raw_name is not None else ""

        if not item_id:
            raise ValueError(f"Row {row_idx}: id is empty")
        if not name:
            raise ValueError(f"Row {row_idx}: name is empty")
        if raw_chance is None or str(raw_chance).strip() == "":
            raise ValueError(f"Row {row_idx}: chance is empty")

        try:
            weight = float(raw_chance)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Row {row_idx}: chance must be float") from exc

        if not (0 < weight < 1):
            raise ValueError(f"Row {row_idx}: chance must satisfy 0 < chance < 1")

        if item_id in seen_ids:
            raise ValueError(f"Row {row_idx}: duplicate id '{item_id}'")
        seen_ids.add(item_id)

        icon_path = icons_dir / f"{item_id}.png"
        if not icon_path.exists():
            raise ValueError(f"Row {row_idx}: icon missing '{icon_path}'")

        items.append(Item(item_id=item_id, name=name, weight=weight, icon_path=icon_path))

    if not items:
        raise ValueError("No items found in Excel")

    prefix_sums: List[float] = []
    running = 0.0
    for item in items:
        running += item.weight
        prefix_sums.append(running)

    return LootTable(items=items, prefix_sums=prefix_sums, sum_weights=running)
